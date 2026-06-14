"""
Screen: Build Metadata
Generate, view, edit, and track history for one database at a time.
This is the primary working screen — all actions flow through here.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st

from config.strings import UILabels, UIMessages

logger = logging.getLogger(__name__)

# Columns written to / read from the Excel file
_EXPORT_HEADERS = [
    "source_db", "schema_name", "table_name", "column_name",
    "data_type", "description", "human_notes", "guardian_status",
]
# Columns the user is allowed to edit (rest are identifiers / read-only)
_EDITABLE_COLS = {"description", "human_notes"}
# schema_name is validated if present but not required (backwards compat with older exports)
_REQUIRED_IMPORT_COLS = {"source_db", "table_name", "column_name", "description", "human_notes"}


def render() -> None:
    """Render the Build Metadata screen."""
    st.title(UILabels.PAGE_BUILD_TITLE)
    st.write(UILabels.PAGE_BUILD_DESCRIPTION)

    _initialise_session_state()

    # ── Database selector ────────────────────────────────────────────────────
    from database.connection_registry import get_registry
    registry = get_registry()
    db_names = registry.all_names()

    if not db_names:
        st.warning(UILabels.NO_DATABASES_CONFIGURED)
        st.info("Add an entry to `config/databases.json` and a matching connection string to `.env`, then restart.")
        return

    col_select, col_generate, col_export, col_import_lbl = st.columns([3, 1, 1, 1])

    with col_select:
        selected_db = st.selectbox(
            UILabels.LABEL_SELECT_DB,
            options=db_names,
            index=db_names.index(st.session_state.build_selected_db)
            if st.session_state.build_selected_db in db_names
            else 0,
            key="build_db_selectbox",
        )
        st.session_state.build_selected_db = selected_db

    with col_generate:
        st.markdown(" ")
        generate_clicked = st.button(
            UILabels.BTN_GENERATE,
            type="primary",
            use_container_width=True,
        )

    # ── Load entries once — shared by export button and table cards ──────────
    from vector_store.qdrant_store import get_store
    store = get_store()
    all_entries = _load_entries(store, selected_db)

    with col_export:
        st.markdown(" ")
        if all_entries:
            excel_bytes = _build_excel(selected_db, all_entries)
            st.download_button(
                "⬇ Export",
                data=excel_bytes,
                file_name=f"{selected_db}_metadata.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.button("⬇ Export", disabled=True, use_container_width=True)

    with col_import_lbl:
        st.markdown(" ")
        import_clicked = st.button("⬆ Import", use_container_width=True)

    # ── Import panel (shown when Import button clicked) ──────────────────────
    if import_clicked:
        st.session_state["show_import"] = True
    if st.session_state.get("show_import"):
        _render_import_panel(selected_db, all_entries)

    # ── Generation ──────────────────────────────────────────────────────────
    if generate_clicked:
        _run_generation(selected_db)

    # ── Table cards ─────────────────────────────────────────────────────────
    _render_tables_view_from_entries(all_entries, selected_db)

    # ── History section ──────────────────────────────────────────────────────
    st.divider()
    _render_history_section(selected_db)


# ── Initialisation ────────────────────────────────────────────────────────────


def _initialise_session_state() -> None:
    """Set up session state keys on first load."""
    defaults = {
        "build_selected_db": None,
        "show_import": False,
    }
    for key, default in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default


# ── Entry loader ──────────────────────────────────────────────────────────────


def _load_entries(store: Any, database_name: str) -> List[Dict[str, Any]]:
    """Load and deduplicate all entries for database_name from Qdrant."""
    raw = store.get_all_entries(filters={"source_db": database_name}, limit=5000)
    seen: set = set()
    unique: List[Dict[str, Any]] = []
    for e in raw:
        eid = e.get("id", "")
        if eid not in seen:
            seen.add(eid)
            unique.append(e)
    return unique


# ── Export ────────────────────────────────────────────────────────────────────


def _build_excel(database_name: str, entries: List[Dict[str, Any]]) -> bytes:
    """Build an Excel workbook from the current metadata entries."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = database_name[:31]  # Excel sheet name limit

    # Header row
    ws.append(_EXPORT_HEADERS)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Identifier column fill (light gray — signals read-only)
    id_fill = PatternFill("solid", fgColor="F2F2F2")

    sorted_entries = sorted(
        entries,
        key=lambda e: (e.get("table_name", "").lower(), e.get("column_name", "").lower()),
    )
    for entry in sorted_entries:
        row = [
            entry.get("source_db", ""),
            entry.get("schema_name", ""),
            entry.get("table_name", ""),
            entry.get("column_name", ""),
            entry.get("data_type", ""),
            entry.get("description", ""),
            entry.get("human_notes", ""),
            entry.get("guardian_status", ""),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx, header in enumerate(_EXPORT_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if header not in _EDITABLE_COLS:
                cell.fill = id_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = {
        "source_db": 20, "schema_name": 12, "table_name": 25, "column_name": 25,
        "data_type": 15, "description": 60, "human_notes": 45, "guardian_status": 16,
    }
    for col_idx, header in enumerate(_EXPORT_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(header, 20)

    ws.freeze_panes = "F2"  # freeze source_db + schema + table + column + data_type; keep description visible

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Import ────────────────────────────────────────────────────────────────────


def _render_import_panel(database_name: str, existing_entries: List[Dict[str, Any]]) -> None:
    """Render the file uploader and process an uploaded Excel."""
    with st.container(border=True):
        st.markdown("**Import metadata from Excel**")
        st.caption(
            "Upload the exported Excel file after editing `description` and `human_notes` columns. "
            "Other columns are ignored. The `source_db` must match the selected database."
        )
        uploaded = st.file_uploader(
            "Choose .xlsx file",
            type=["xlsx"],
            key="import_file_uploader",
        )
        col_ok, col_cancel = st.columns([1, 5])
        with col_cancel:
            if st.button("✕ Close", key="import_close"):
                st.session_state["show_import"] = False
                st.rerun()
        if uploaded is not None:
            with col_ok:
                if st.button("Apply import", type="primary", key="import_apply"):
                    _process_import(database_name, uploaded.read(), existing_entries)


def _process_import(
    database_name: str,
    file_bytes: bytes,
    existing_entries: List[Dict[str, Any]],
) -> None:
    """Validate and apply bulk updates from the uploaded Excel."""
    from openpyxl import load_workbook
    from pipeline.single_db_pipeline import save_user_edit

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        st.error(f"Could not open file: {exc}")
        return

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        st.error("The file is empty.")
        return

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = _REQUIRED_IMPORT_COLS - set(headers)
    if missing:
        st.error(f"Missing required columns: {sorted(missing)}. Expected: {sorted(_REQUIRED_IMPORT_COLS)}")
        return

    col = {h: i for i, h in enumerate(headers)}

    # Resolve expected schema for this database (used when schema_name column is present)
    from database.connection_registry import get_registry
    db_config = get_registry().get_config(database_name)
    expected_schema = db_config.schema.lower() if db_config and db_config.schema else None
    has_schema_col = "schema_name" in col

    # Build lookup: (schema.lower, table.lower, column.lower) → entry_id
    lookup: Dict[Tuple[str, str, str], str] = {
        (
            e.get("schema_name", "").lower(),
            e.get("table_name", "").lower(),
            e.get("column_name", "").lower(),
        ): e.get("id", "")
        for e in existing_entries
        if e.get("id")
    }

    saved = errors = 0
    error_msgs: List[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        source_db = str(row[col["source_db"]] or "").strip()
        table_name = str(row[col["table_name"]] or "").strip()
        column_name = str(row[col["column_name"]] or "").strip()
        description = str(row[col["description"]] or "").strip()
        human_notes = str(row[col["human_notes"]] or "").strip()

        # Skip blank rows
        if not table_name and not description:
            continue

        # Validate source_db
        if source_db.lower() != database_name.lower():
            errors += 1
            error_msgs.append(f"Row {row_num}: source_db '{source_db}' does not match selected database '{database_name}'")
            continue

        # Validate schema_name if the column exists in the file
        if has_schema_col and expected_schema:
            file_schema = str(row[col["schema_name"]] or "").strip().lower()
            if file_schema and file_schema != expected_schema:
                errors += 1
                error_msgs.append(f"Row {row_num}: schema_name '{file_schema}' does not match expected schema '{expected_schema}' for '{database_name}'")
                continue

        schema_in_file = str(row[col["schema_name"]] if has_schema_col and col.get("schema_name") is not None else "").strip().lower()
        # Validate table/column exists — match on schema + table + column
        key = (schema_in_file, table_name.lower(), column_name.lower())
        entry_id = lookup.get(key)
        # Fallback: match without schema for files exported before schema was added
        if not entry_id:
            key_no_schema = ("", table_name.lower(), column_name.lower())
            entry_id = lookup.get(key_no_schema)
        if not entry_id:
            col_label = f"'{table_name}' / '{column_name}'" if column_name else f"'{table_name}' (table-level)"
            errors += 1
            error_msgs.append(f"Row {row_num}: {col_label} not found in '{database_name}' — check spelling")
            continue

        ok = save_user_edit(entry_id, description, human_notes)
        if ok:
            saved += 1
        else:
            errors += 1
            error_msgs.append(f"Row {row_num}: save failed for {table_name}.{column_name}")

    if errors:
        st.warning(f"Import done — **{saved} saved**, **{errors} failed**.")
        with st.expander(f"Show {errors} error(s)"):
            for msg in error_msgs:
                st.error(msg)
    else:
        st.success(f"Import complete — {saved} entries updated.")
        st.session_state["show_import"] = False
        st.rerun()


# ── Generation ────────────────────────────────────────────────────────────────


def _run_generation(database_name: str) -> None:
    """
    Run the single-database pipeline inline using st.status so the user sees
    table-by-table progress as each LLM call completes.
    """
    from pipeline.single_db_pipeline import process_database

    generated_tables: List[str] = []
    errors: List[str] = []

    with st.status(
        f"Generating metadata for **{database_name}**...",
        expanded=True,
    ) as status:

        def on_progress(event: str, table_name: str, detail: str) -> None:
            if event == "start":
                status.write(f"⏳ {table_name} — {detail}")
            elif event == "done":
                status.write(f"✅ {table_name} — {detail}")
                generated_tables.append(table_name)
            elif event == "error":
                status.write(f"❌ {table_name} — {detail}")
                errors.append(f"{table_name}: {detail}")

        process_database(database_name, progress_fn=on_progress)

        if errors:
            status.update(
                label=f"Finished with {len(errors)} error(s). {len(generated_tables)} tables stored.",
                state="error",
            )
        else:
            status.update(
                label=UILabels.GENERATION_COMPLETE.format(table_count=len(generated_tables)),
                state="complete",
            )

    if errors:
        for err in errors:
            st.error(err)

    st.rerun()


# ── Table cards ───────────────────────────────────────────────────────────────


def _render_tables_view_from_entries(
    unique_entries: List[Dict[str, Any]],
    database_name: str,
) -> None:
    """Render one card per table using the already-loaded, deduplicated entries."""
    if not unique_entries:
        st.info(UILabels.NO_ENTRIES_FOR_DB)
        return

    # Group by lowercased table_name so mixed-case DB names don't create duplicate cards
    tables: Dict[str, List[Dict[str, Any]]] = {}
    display_names: Dict[str, str] = {}
    for entry in unique_entries:
        raw_name = entry.get("table_name", "")
        key = raw_name.lower()
        display_names.setdefault(key, raw_name)
        tables.setdefault(key, []).append(entry)

    table_keys = sorted(tables.keys())
    st.markdown(f"**{len(table_keys)} tables** in `{database_name}`")

    for table_key in table_keys:
        _render_table_card(display_names[table_key], tables[table_key], database_name)


def _render_table_card(
    table_name: str,
    entries: List[Dict[str, Any]],
    database_name: str,
) -> None:
    """Render one expandable card for a table with editable descriptions and save button."""
    table_entry = next((e for e in entries if not e.get("column_name")), None)
    column_entries = sorted(
        [e for e in entries if e.get("column_name")],
        key=lambda e: e.get("column_name", ""),
    )

    status = (table_entry or (entries[0] if entries else {})).get("guardian_status", "pending")
    human_verified = (table_entry or (entries[0] if entries else {})).get("human_verified", False)

    badge = {
        "approved": "🟢",
        "needs_review": "🟡",
        "rejected": "🔴",
        "pending": "⚪",
    }.get(status, "⚪")
    verified_tag = " ✅ Human verified" if human_verified else ""

    with st.expander(f"{badge} **{table_name}**{verified_tag} — {len(column_entries)} columns"):

        # ── Table-level description ─────────────────────────────────────────
        if table_entry:
            st.markdown(f"**{UILabels.LABEL_TABLE_DESCRIPTION}**")
            table_desc_key = f"desc_{table_entry['id']}"
            table_notes_key = f"notes_{table_entry['id']}"
            if table_desc_key not in st.session_state:
                st.session_state[table_desc_key] = table_entry.get("description", "")
            if table_notes_key not in st.session_state:
                st.session_state[table_notes_key] = table_entry.get("human_notes", "")

            st.text_area(
                "Table-level description",
                key=table_desc_key,
                height=100,
                label_visibility="collapsed",
            )
            st.text_input("Notes (optional)", key=table_notes_key)

        # ── Column descriptions ─────────────────────────────────────────────
        if column_entries:
            st.markdown("---")
            st.markdown("**Columns**")
            for col_entry in column_entries:
                col_name = col_entry.get("column_name", "")
                col_type = col_entry.get("data_type", "")
                col_status = col_entry.get("guardian_status", "pending")
                col_badge = {"approved": "🟢", "needs_review": "🟡", "rejected": "🔴", "pending": "⚪"}.get(col_status, "⚪")

                st.markdown(f"{col_badge} `{col_name}` — *{col_type}*")

                desc_key = f"desc_{col_entry['id']}"
                notes_key = f"notes_{col_entry['id']}"
                if desc_key not in st.session_state:
                    st.session_state[desc_key] = col_entry.get("description", "")
                if notes_key not in st.session_state:
                    st.session_state[notes_key] = col_entry.get("human_notes", "")

                st.text_area(
                    f"Description for {col_name}",
                    key=desc_key,
                    height=80,
                    label_visibility="collapsed",
                )
                st.text_input(
                    f"Notes for {col_name}",
                    key=notes_key,
                    label_visibility="collapsed",
                    placeholder="Add a business note (optional)",
                )
                st.markdown("")

        # ── Action buttons ──────────────────────────────────────────────────
        btn_col_save, btn_col_regen, _ = st.columns([1, 1, 2])

        with btn_col_save:
            if st.button(
                UILabels.BTN_SAVE_TABLE,
                key=f"save_{table_name}_{database_name}",
                type="primary",
            ):
                _save_table_edits(table_entry, column_entries)

        with btn_col_regen:
            if st.button(
                UILabels.BTN_REGENERATE_TABLE,
                key=f"regen_{table_name}_{database_name}",
            ):
                _regenerate_table(database_name, table_name)


def _save_table_edits(
    table_entry: Optional[Dict[str, Any]],
    column_entries: List[Dict[str, Any]],
) -> None:
    """Save all edited descriptions for a table and its columns."""
    from pipeline.single_db_pipeline import save_user_edit

    saved = 0
    errors = 0

    all_entries = ([table_entry] if table_entry else []) + column_entries

    for entry in all_entries:
        entry_id = entry.get("id", "")
        new_desc = st.session_state.get(f"desc_{entry_id}", entry.get("description", ""))
        new_notes = st.session_state.get(f"notes_{entry_id}", entry.get("human_notes", ""))

        success = save_user_edit(entry_id, new_desc, new_notes)
        if success:
            saved += 1
        else:
            errors += 1

    if errors:
        st.error(f"Saved {saved} entries. {errors} failed.")
    else:
        st.success(f"Saved {saved} entries. Changes recorded in history.")
        st.rerun()


def _regenerate_table(database_name: str, table_name: str) -> None:
    """Re-run LLM enrichment for a single table."""
    from pipeline.single_db_pipeline import process_single_table_by_name

    with st.spinner(f"Re-generating metadata for {table_name}..."):
        entries = process_single_table_by_name(database_name, table_name)

    if entries:
        st.success(f"Re-generated {len(entries)} entries for {table_name}.")
        st.rerun()
    else:
        st.error(f"Re-generation failed for {table_name}. Check logs.")


# ── History section ───────────────────────────────────────────────────────────


def _render_history_section(database_name: str) -> None:
    """Show all archived versions for the selected database at the bottom of the screen."""
    from database.sqlite_store import get_sqlite_store

    sqlite = get_sqlite_store()

    with st.expander(f"📋 {UILabels.LABEL_HISTORY}", expanded=False):
        history = sqlite.get_entry_history(source_db=database_name, limit=300)

        if not history:
            st.info(UILabels.HISTORY_EMPTY)
            return

        # Table filter
        all_tables = sorted({h["table_name"] for h in history if h.get("table_name")})
        filter_table = st.selectbox(
            UILabels.LABEL_FILTER_TABLE_HISTORY,
            options=["All"] + all_tables,
            key=f"history_filter_{database_name}",
        )
        if filter_table != "All":
            history = [h for h in history if h.get("table_name") == filter_table]

        # Export button
        csv_bytes = _build_history_csv(history)
        st.download_button(
            UILabels.BTN_EXPORT_HISTORY,
            data=csv_bytes,
            file_name=f"history_{database_name}.csv",
            mime="text/csv",
        )

        st.markdown(f"**{len(history)} version records**")

        # Render as a table
        rows = []
        for h in history:
            changed_at = h["changed_at"]
            if hasattr(changed_at, "strftime"):
                changed_at_str = changed_at.strftime("%Y-%m-%d %H:%M")
            else:
                changed_at_str = str(changed_at)[:16]

            col_label = h.get("column_name") or "(table)"
            desc_preview = (h.get("description") or "")[:80]
            if len(h.get("description") or "") > 80:
                desc_preview += "…"

            rows.append({
                "Table": h.get("table_name", ""),
                "Column": col_label,
                "Ver": h.get("version", ""),
                "Changed At": changed_at_str,
                "By": h.get("changed_by", ""),
                "Type": h.get("change_type", ""),
                "Status": h.get("guardian_status", ""),
                "Description preview": desc_preview,
            })

        if rows:
            st.dataframe(rows, use_container_width=True, hide_index=True)


def _build_history_csv(history: List[Dict[str, Any]]) -> bytes:
    """Build a full CSV export of the history records."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "entry_id", "table_name", "column_name", "version",
        "changed_at", "changed_by", "change_type",
        "guardian_status", "description", "human_notes",
    ])
    for h in history:
        changed_at = h.get("changed_at", "")
        if hasattr(changed_at, "isoformat"):
            changed_at = changed_at.isoformat()
        writer.writerow([
            h.get("entry_id", ""),
            h.get("table_name", ""),
            h.get("column_name", ""),
            h.get("version", ""),
            changed_at,
            h.get("changed_by", ""),
            h.get("change_type", ""),
            h.get("guardian_status", ""),
            h.get("description", ""),
            h.get("human_notes", ""),
        ])
    return output.getvalue().encode("utf-8")
